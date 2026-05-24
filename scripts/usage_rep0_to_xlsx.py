"""Build an Excel summary of rep_0 usage.json data per regime.

Walks runs/sweep_l1_048/regime_*/rep_0/**/usage.json, merges
``coder`` + ``coder-translate`` into a single ``Coder`` bucket so each
row reports planner / coder / reviewer / total, and writes
``usage_rep0_summary.xlsx`` with a two-row header (merged metric cell
spanning four sub-columns: P / C / R / Total).
"""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

REPO_ROOT = Path(__file__).resolve().parent.parent
SWEEP_DIR = REPO_ROOT / "runs" / "sweep_l1_048"

REGIMES = [
    ("01", "regime_01_minimal"),
    ("02", "regime_02_lean_balanced"),
    ("03", "regime_03_default"),
    ("04", "regime_04_coder_wide"),
    ("05", "regime_05_reviewer_deep"),
    ("06", "regime_06_patient_coder"),
    ("07", "regime_07_maxed"),
    ("08", "regime_08_default_depth_3"),
    ("09", "regime_09_default_depth_6"),
    ("10", "regime_10_default_depth_15"),
]

METRICS = [
    ("invocations", "Invocations"),
    ("turns", "Turns"),
    ("input_tokens", "Input tokens"),
    ("output_tokens", "Output tokens"),
    ("reasoning_output_tokens", "Reasoning output tokens"),
]

AGENT_LABELS = ["Planner", "Coder", "Reviewer", "Total"]

_SOL_RE = re.compile(r"SOL score:\s+([0-9.]+)")
_BEST_RE = re.compile(r"Best:\s+([0-9.]+)\s*us")


def load_rep0(name: str) -> dict | None:
    rep0 = SWEEP_DIR / name / "rep_0"
    if not rep0.exists():
        return None
    found = list(rep0.rglob("usage.json"))
    if not found:
        return None
    with found[0].open() as f:
        data = json.load(f)
    by_agent = data["by_agent"]
    p = by_agent["planner"]
    c = by_agent["coder"]
    ct = by_agent["coder-translate"]
    r = by_agent["reviewer"]
    coder_merged = {k: c[k] + ct[k] for k in c}

    # report.txt sits beside usage.json in the same run_<ts>/ dir.
    report_path = found[0].parent / "report.txt"
    sol_score: float | None = None
    runtime_us: float | None = None
    if report_path.exists():
        text = report_path.read_text()
        sol_match = _SOL_RE.search(text)
        best_match = _BEST_RE.search(text)
        sol_score = float(sol_match.group(1)) if sol_match else None
        runtime_us = float(best_match.group(1)) if best_match else None

    return {
        "planner": p,
        "coder": coder_merged,
        "reviewer": r,
        "total": data["total"],
        "sol_score": sol_score,
        "runtime_us": runtime_us,
    }


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "rep_0 usage"

    header_fill = PatternFill("solid", fgColor="DDDDDD")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Two-row header: row 1 = metric group (merged across 4 sub-cols),
    # row 2 = sub-column labels (Planner / Coder / Reviewer / Total).
    ws.cell(row=1, column=1, value="Case").font = bold
    ws.cell(row=1, column=1).alignment = center
    ws.cell(row=1, column=1).fill = header_fill
    ws.cell(row=1, column=1).border = border
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=2, column=1).border = border

    col = 2
    for _, label in METRICS:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
        top = ws.cell(row=1, column=col, value=label)
        top.font = bold
        top.alignment = center
        top.fill = header_fill
        for j in range(4):
            top_cell = ws.cell(row=1, column=col + j)
            top_cell.border = border
            sub = ws.cell(row=2, column=col + j, value=AGENT_LABELS[j])
            sub.font = bold
            sub.alignment = center
            sub.fill = header_fill
            sub.border = border
        col += 4

    # Scalar trailing columns: SOL score + best runtime (best-of-tree at end).
    scalar_labels = ["SOL score", "Runtime (µs)"]
    scalar_start_col = col
    for label in scalar_labels:
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = bold
        cell.alignment = center
        cell.fill = header_fill
        cell.border = border
        ws.cell(row=2, column=col).border = border
        col += 1

    # Data rows.
    row = 3
    for case, name in REGIMES:
        agents = load_rep0(name)
        ws.cell(row=row, column=1, value=case).alignment = center
        ws.cell(row=row, column=1).border = border

        col = 2
        for key, _ in METRICS:
            if agents is None:
                for j in range(4):
                    cell = ws.cell(row=row, column=col + j, value="—")
                    cell.alignment = center
                    cell.border = border
                col += 4
                continue

            total = agents["total"][key]
            p_v = agents["planner"][key]
            c_v = agents["coder"][key]
            r_v = agents["reviewer"][key]

            def pct(v: int) -> float:
                return (100.0 * v / total) if total else 0.0

            cells = [
                f"{p_v} ({pct(p_v):.1f}%)",
                f"{c_v} ({pct(c_v):.1f}%)",
                f"{r_v} ({pct(r_v):.1f}%)",
                str(total),
            ]
            for j, val in enumerate(cells):
                cell = ws.cell(row=row, column=col + j, value=val)
                cell.alignment = center
                cell.border = border
            col += 4

        # Trailing scalar columns.
        sol = agents.get("sol_score") if agents else None
        runtime = agents.get("runtime_us") if agents else None
        sol_cell = ws.cell(
            row=row, column=col,
            value="—" if sol is None else round(sol, 4),
        )
        sol_cell.alignment = center
        sol_cell.border = border
        rt_cell = ws.cell(
            row=row, column=col + 1,
            value="—" if runtime is None else round(runtime, 2),
        )
        rt_cell.alignment = center
        rt_cell.border = border
        row += 1

    # Column widths.
    from openpyxl.utils import get_column_letter

    ws.column_dimensions["A"].width = 8
    metric_end = 2 + 4 * len(METRICS)
    for letter_idx in range(2, metric_end):
        ws.column_dimensions[get_column_letter(letter_idx)].width = 16
    for letter_idx in range(metric_end, metric_end + 2):
        ws.column_dimensions[get_column_letter(letter_idx)].width = 14
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22

    ws.freeze_panes = "B3"

    out = SWEEP_DIR / "usage_rep0_summary.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
